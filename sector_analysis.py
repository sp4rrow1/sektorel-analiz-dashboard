# -*- coding: utf-8 -*-
"""
Sektorel LLM Analiz Modulu — Kimteks Word formati icin yeniden yazildi.
LLM'den [TAG] bolumlu yapi beklenir; report_word.py parse eder.
"""
import re
from llm_client import call_llm
from nace_config import SECTOR_NAMES, SITC_MAP

_TR_AY = ['Ocak', 'Şubat', 'Mart', 'Nisan', 'Mayıs', 'Haziran',
          'Temmuz', 'Ağustos', 'Eylül', 'Ekim', 'Kasım', 'Aralık']

def _tr_num(v, dec=1):
    """1234.5 -> '1.234,5' (Türkçe biçim)."""
    if v is None:
        return "—"
    s = f"{v:,.{dec}f}"
    return s.replace(",", "\x00").replace(".", ",").replace("\x00", ".")

def _tr_donem(p):
    try:
        yr, mo = p.split("-")
        return f"{_TR_AY[int(mo)-1]} {yr}"
    except Exception:
        return p or ""

def _yon(v):
    return "artış" if (v or 0) >= 0 else "azalış"

def _yon2(v):
    return "arttığı" if (v or 0) >= 0 else "azaldığı"

def _first_ser(fd):
    if not fd: return {}
    return dict(sorted(next(iter(fd.values())).items()))

def _merged(fd):
    m = {}
    for _, s in (fd or {}).items():
        for p, v in s.items():
            if v is not None: m.setdefault(p, []).append(v)
    return {p: sum(vs)/len(vs) for p, vs in m.items()}

def _last(s):
    pts = [(p, v) for p, v in sorted(s.items()) if v is not None]
    return pts[-1] if pts else (None, None)

def _yil_ort(s, yr):
    vs = [v for p, v in s.items() if v is not None and p.startswith(str(yr))]
    return sum(vs)/len(vs) if vs else None


def fallback_report(nace, fig1, fig2, fig3, fig4, fig5, iso_agg=None, fig6=None, fig7=None):
    """
    LLM erişilemediğinde gerçek verilerden deterministik, banka üslubunda
    [TAG] bölümlü rapor üretir. Böylece rapor asla boş kalmaz.
    """
    sector = SECTOR_NAMES.get(nace, nace)
    try:
        from report_charts import nace_name as _nn
        official = _nn(nace)
        if official and official != nace:
            sector = official
    except Exception:
        pass

    prod = _merged(fig1)
    pp, pv = _last(prod)
    son_yil = int(pp.split("-")[0]) if pp else None
    prev_yil = (son_yil - 1) if son_yil else None
    prod_yil = _yil_ort(prod, prev_yil) if prev_yil else None

    out = []
    out.append("[GIRIS]")
    out.append(
        f"Türkiye'de {sector.lower()} sektörü, imalat sanayiinin önemli bir bileşenini "
        f"oluşturmakta ve ülke ekonomisinde üretim, istihdam ve dış ticaret açısından "
        f"belirleyici bir rol üstlenmektedir. Sektörün güncel görünümüne ilişkin temel "
        f"göstergeler ve değerlendirmelerimiz aşağıda yer almaktadır.")

    # ── SEKIL1: Üretim ──
    out.append("[SEKIL1]")
    if pv is not None:
        s = ""
        if prod_yil is not None:
            s += (f"{sector} sektörü üretimi {prev_yil} yılında bir önceki yıla göre "
                  f"%{_tr_num(abs(prod_yil))} oranında {_yon(prod_yil)} kaydetmiştir. ")
        s += (f"{_tr_donem(pp)} itibarıyla üretimin, geçen yılın aynı ayına göre "
              f"%{_tr_num(abs(pv))} {_yon2(pv)} gözlemlenmektedir.")
        out.append(s)
    else:
        out.append(f"{sector} sektörüne ilişkin üretim endeksi verisi bu dönem için sınırlıdır.")

    # ── SEKIL2: Alt kırılımlar ──
    out.append("[SEKIL2]")
    if fig2:
        lasts = {}
        for code, s in fig2.items():
            _, v = _last(s)
            if v is not None: lasts[code] = v
        if lasts:
            en_iyi = max(lasts, key=lasts.get)
            en_kotu = min(lasts, key=lasts.get)
            def _cn(c):
                try:
                    from report_charts import nace_name as _nn
                    return _nn(c)
                except Exception:
                    return c
            s = (f"Alt sektörler incelendiğinde, son dönemde en yüksek {_yon(lasts[en_iyi])} "
                 f"%{_tr_num(lasts[en_iyi])} ile \"{_cn(en_iyi)}\" alt sektöründe gerçekleşmiştir.")
            if en_kotu != en_iyi:
                s += (f" Buna karşılık, \"{_cn(en_kotu)}\" alt sektöründe %{_tr_num(abs(lasts[en_kotu]))} "
                      f"oranında {_yon(lasts[en_kotu])} kaydedilmiştir.")
            out.append(s)
        else:
            out.append("Alt sektör kırılımlarına ilişkin veri bu dönem için sınırlıdır.")
    else:
        out.append("Alt sektör kırılımlarına ilişkin veri bu dönem için mevcut değildir.")

    # ── SEKIL3: Dış ticaret ──
    out.append("[SEKIL3]")
    ih = {k: v for k, v in (fig3 or {}).items() if "hracat" in k and "thalat" not in k}
    it = {k: v for k, v in (fig3 or {}).items() if "thalat" in k}
    ihp, ihv = _last(_first_ser(ih)); itp, itv = _last(_first_ser(it))
    if ihv is not None or itv is not None:
        don = _tr_donem(ihp or itp)
        parts = []
        if ihv is not None:
            parts.append(f"ihracatın %{_tr_num(abs(ihv))} oranında {_yon2(ihv)}")
        if itv is not None:
            parts.append(f"ithalatın ise %{_tr_num(abs(itv))} oranında {_yon2(itv)}")
        out.append(f"{don} döneminde, önceki yılın aynı dönemine göre {', '.join(parts)} görülmektedir.")
    else:
        out.append(f"{sector} sektörüne ilişkin dış ticaret verisi bu dönem için sınırlıdır.")

    # ── SEKIL4: KKO ──
    out.append("[SEKIL4]")
    kko_sec = {k: v for k, v in (fig4 or {}).items() if "anayii" not in k}
    kko_im  = {k: v for k, v in (fig4 or {}).items() if "anayii" in k}
    ksp, ksv = _last(_first_ser({k: (dict(v) if not isinstance(v, dict) else v) for k, v in kko_sec.items()}))
    kip, kiv = _last(_first_ser({k: (dict(v) if not isinstance(v, dict) else v) for k, v in kko_im.items()}))
    if ksv is not None:
        s = (f"{_tr_donem(ksp)} itibarıyla kapasite kullanım oranı, {sector} sektöründe "
             f"%{_tr_num(ksv)} düzeyinde gerçekleşmiştir.")
        if kiv is not None:
            fark = ksv - kiv
            s += (f" Aynı dönemde imalat sanayii genelinde %{_tr_num(kiv)} olarak ölçülen oran, "
                  f"sektörün imalat ortalamasının {_tr_num(abs(fark))} puan "
                  f"{'üzerinde' if fark >= 0 else 'altında'} seyrettiğine işaret etmektedir.")
        out.append(s)
    else:
        out.append("Kapasite kullanım oranına ilişkin veri bu dönem için sınırlıdır.")

    # ── SEKIL5: ÜFE + reel ciro ──
    out.append("[SEKIL5]")
    ufe = _first_ser(fig5)
    ufe_im = {k: v for k, v in (fig5 or {}).items() if " c " in (" " + k.lower() + " ")}
    usp, usv = _last(ufe); uip, uiv = _last(_first_ser(ufe_im))
    if usv is not None:
        s = (f"{sector} sektörü üretici fiyatları, {_tr_donem(usp)} itibarıyla yıllık "
             f"%{_tr_num(usv)} düzeyinde değişim göstermiştir.")
        if uiv is not None:
            s += (f" Aynı dönemde imalat sanayii ÜFE değişimi %{_tr_num(uiv)} olup, sektör fiyatları "
                  f"imalat geneli ile {'paralel' if abs(usv-uiv) < 5 else 'kısmen ayrışan'} bir seyir izlemektedir.")
        # reel ciro
        ciro = _first_ser(fig6); cp, cv = _last(ciro)
        if cv is not None and usv is not None:
            reel = ((1 + cv/100.0) / (1 + usv/100.0) - 1) * 100.0
            s += (f" Nominal ciro değişiminin (%{_tr_num(cv)}) enflasyondan arındırılmış reel "
                  f"karşılığı %{_tr_num(reel)} olarak hesaplanmakta; bu durum sektörde reel "
                  f"{'büyümeye' if reel >= 0 else 'daralmaya'} işaret etmektedir.")
        out.append(s)
    else:
        out.append("Üretici fiyat endeksine ilişkin veri bu dönem için sınırlıdır.")

    # ── SEKIL6: İSO ──
    if iso_agg:
        out.append("[SEKIL6]")
        yil = iso_agg.get('yil', '')
        s = (f"Türkiye'nin en büyük sanayi kuruluşlarını kapsayan İSO 500 ve İkinci 500 "
             f"listelerinde {sector} sektöründen {yil} yılında {iso_agg['firma_sayisi']} kuruluş "
             f"yer almıştır. Bu kuruluşların toplam üretimden satışları "
             f"{_tr_num(iso_agg['toplam_uretim_satis']/1e9)} milyar TL, toplam ihracatı ise "
             f"{_tr_num(iso_agg['toplam_ihracat_musd'], 0)} milyon dolar düzeyinde gerçekleşmiştir.")
        if iso_agg.get('favok_marj_med') is not None:
            s += f" Sektörün medyan FAVÖK marjı %{_tr_num(iso_agg['favok_marj_med'])} olarak hesaplanmıştır."
        out.append(s)

    return "\n".join(out)


def summarize_data(nace, fig1, fig2, fig3, fig4, fig5):
    """5 sekildeki veriyi LLM icin ozet metin haline getirir."""
    def last_n(data_dict, n=12):
        out = {}
        for lbl, series in (data_dict or {}).items():
            pts = [(p, v) for p, v in sorted(series.items()) if v is not None][-n:]
            out[lbl] = pts
        return out

    def fmt_son(pts):
        """Son nokta metni — veri yoksa guvenli."""
        if not pts:
            return 'veri yok'
        p, v = pts[-1]
        return f'{v:.1f}% ({p})'

    def trend(pts):
        if len(pts) < 2: return 'belirsiz'
        vals = [v for _, v in pts if v is not None]
        if not vals: return 'veri yok'
        a = sum(vals[:3]) / max(len(vals[:3]), 1)
        b = sum(vals[-3:]) / max(len(vals[-3:]), 1)
        diff = b - a
        if diff > 3:  return 'yukari yonlu'
        if diff < -3: return 'asagi yonlu'
        return 'yatay seyir'

    def fmt(pts):
        return ', '.join(f'{p}: {v:.1f}%' for p, v in pts[-6:] if v is not None)

    lines = []
    sector = SECTOR_NAMES.get(nace, nace)
    lines.append(f'SEKTOR: {nace} - {sector}')
    lines.append(f'SITC eslesmesi: {SITC_MAP.get(nace, "T")}\n')

    if fig1:
        lines.append('=== SEKIL 1: Uretim Endeksi (YoY %) ===')
        for lbl, pts in last_n(fig1, 12).items():
            lines.append(f'  {lbl}: son={fmt_son(pts)}, trend={trend(pts)}')
            lines.append(f'  Son 6 ay: {fmt(pts)}')

    if fig2:
        lines.append('\n=== SEKIL 2: Alt Kirilimlar (YoY %) ===')
        for lbl, pts in last_n(fig2, 12).items():
            lines.append(f'  {lbl}: son={fmt_son(pts)}, trend={trend(pts)}')

    if fig3:
        lines.append('\n=== SEKIL 3: Dis Ticaret (YoY %) ===')
        for lbl, pts in last_n(fig3, 12).items():
            lines.append(f'  {lbl}: son={fmt_son(pts)}, trend={trend(pts)}')

    if fig4:
        lines.append('\n=== SEKIL 4: Kapasite Kullanim Orani (%) ===')
        for lbl, v in (fig4 or {}).items():
            d = v if isinstance(v, dict) else dict(v)
            pts_list = [(p, x) for p, x in sorted(d.items()) if x is not None][-12:]
            lines.append(f'  {lbl}: son={fmt_son(pts_list)}, trend={trend(pts_list)}')

    if fig5:
        lines.append('\n=== SEKIL 5: UFE Yillik Degisim (%) ===')
        for lbl, pts in last_n(fig5, 6).items():
            lines.append(f'  {lbl}: son={fmt_son(pts)}')

    return '\n'.join(lines)


SYSTEM_PROMPT = """Sen Türkiye'nin önde gelen bir kalkınma ve yatırım bankasının Sektörel
Araştırmalar biriminde çalışan kıdemli bir ekonomistsin. Bankacılık standardında,
resmi ve akıcı Türkçe ile sektör raporları yazarsın.

YAZIM KURALLARI (kesinlikle uy):
- Madde işareti (-) KULLANMA. Her bölümü AKICI PARAGRAF olarak yaz (2-4 cümle).
- Sayıları Türkçe biçimde yaz: ondalık ayırıcı VİRGÜL (%3,4 — %3.4 DEĞİL),
  binlik ayırıcı nokta (17.782). Yüzde işareti sayıdan önce (%12,1).
- Somut değerlere daima atıf yap; hem yıllık (ör. "2025 yılında") hem son ay
  (ör. "2026 Mart ayında") karşılaştırması ver.
- Tipik kalıpları kullan: "... bir önceki yıla göre %X oranında artış/azalış
  kaydetmiştir", "... geçen yılın aynı ayına göre %Y arttığı gözlemlenmektedir",
  "... düzeyinde gerçekleşmiştir", "... paralel seyretmektedir".
- Abartısız, olgusal, rasyonel. Başlık, etiket, markdown yıldızı KULLANMA."""


def _derived_metrics_text(fig1, fig5, fig6, fig7):
    """Reel ciro, verimlilik, momentum gibi türev analist metriklerini metne döker."""
    def first_ser(fd):
        if not fd: return {}
        return dict(sorted(next(iter(fd.values())).items()))
    def merged(fd):
        m = {}
        for _, s in (fd or {}).items():
            for p, v in s.items():
                if v is not None: m.setdefault(p, []).append(v)
        return {p: sum(vs)/len(vs) for p, vs in m.items()}
    def last(s):
        pts = [(p, v) for p, v in sorted(s.items()) if v is not None]
        return pts[-1] if pts else (None, None)

    prod = merged(fig1); ciro = first_ser(fig6)
    ufe  = first_ser(fig5); emp = first_ser(fig7)
    lines = []

    cp, cv = last(ciro); up, uv = last(ufe)
    if cv is not None and uv is not None:
        reel = ((1 + cv/100.0) / (1 + uv/100.0) - 1) * 100.0
        lines.append(f"  Reel ciro (nominal %{cv:.1f} - UFE %{uv:.1f}): %{reel:+.1f} "
                     f"({'reel daralma' if reel < 0 else 'reel buyume'})")
    pp, pv = last(prod); ep, ev = last(emp)
    if pv is not None and ev is not None:
        lines.append(f"  Isgucu verimliligi (uretim %{pv:.1f} - istihdam %{ev:.1f}): "
                     f"%{pv - ev:+.1f} puan")
    if not lines:
        return ''
    return "\n=== TUREV ANALIST METRIKLERI ===\n" + "\n".join(lines) + "\n"


def generate_analysis(nace, fig1, fig2, fig3, fig4, fig5, iso_agg=None,
                      fig6=None, fig7=None, kisa=False):
    """
    LLM'e yapılandırılmış sorgu gönderir.
    Dönen metin [TAG] bölümlerine ayrılmış olacak:
      [GIRIS]  — 1-2 paragraf genel değerlendirme
      [SEKIL1] — üretim endeksi maddeleri
      [SEKIL2] — alt kırılım maddeleri
      [SEKIL3] — dış ticaret maddeleri
      [SEKIL4] — KKO maddeleri
      [SEKIL5] — UFE maddeleri
    fig6/fig7 verilirse reel ciro & verimlilik türev metrikleri prompt'a eklenir.
    kisa=True → her bölümde daha az, öz madde (Kısa Özet stili).
    """
    sector = SECTOR_NAMES.get(nace, nace)
    ozet   = summarize_data(nace, fig1, fig2, fig3, fig4, fig5)
    ozet  += _derived_metrics_text(fig1, fig5, fig6, fig7)

    iso_blok = ''
    if iso_agg:
        try:
            from iso_data import iso_summary_text
            iso_txt = iso_summary_text(iso_agg)
            if iso_txt:
                iso_blok = f"\n{iso_txt}\n"
        except Exception:
            pass

    uzunluk = "1-2 cümlelik kısa" if kisa else "2-4 cümlelik"

    prompt = f"""Aşağıdaki TÜİK, TCMB ve İSO verileriyle {nace} - {sector} sektörü için
banka Sektörel Araştırmalar birimi standardında bir rapor yaz.

{ozet}
{iso_blok}

BİÇİM: Yanıtını SADECE aşağıdaki etiketlerle ver. Her etiketin altına, madde
işareti KULLANMADAN, {uzunluk} AKICI BİR PARAGRAF yaz. Sayılarda ondalık ayırıcı
virgül (%3,4), yüzde işareti önde. Örnek üsluba birebir uy.

[GIRIS]
Sektörün stratejik rolünü ve genel görünümünü tanıtan giriş paragrafı.
(Örnek üslup: "Türkiye'de {sector.lower()} sektörü, ... kritik bir rol
oynamaktadır. Sektörün mevcut durumuna ilişkin analizimiz ve temel göstergeler
aşağıda yer almaktadır.")

[SEKIL1]
Üretim endeksi paragrafı. (Örnek üslup: "... sektörü üretimi 2025 yılında bir
önceki yıla göre %X oranında artış kaydetmiştir. 2026 yılı ... ayında ise üretimin
geçen yılın aynı ayına göre %Y arttığı gözlemlenmektedir.")

[SEKIL2]
Alt kırılımlar paragrafı. En yüksek/düşük artış gösteren alt sektörleri isim ve
yüzdeyle belirt. (Örnek üslup: "Seçili alt sektörler incelendiğinde, 2025 yılında
en yüksek artış %X ile '...' alt sektöründe gerçekleşmiştir.")

[SEKIL3]
Dış ticaret paragrafı. İhracat ve ithalatın yıllık değişimini, mümkünse dolar
büyüklüğünü ve dış ticaret dengesini (fazla/açık) yorumla. (Örnek üslup:
"... 2025 yılında yapılan ihracat %X, ithalat ise %Y oranında artış göstermiştir.
2026 ... döneminde ihracatın %Z arttığı, ithalatın ise %W azaldığı görülmektedir.")

[SEKIL4]
Kapasite kullanım oranı paragrafı. Sektörü imalat sanayii ortalamasıyla ve tarihsel
ortalamayla karşılaştır. (Örnek üslup: "... döneminde ortalama kapasite kullanım
oranı imalat sanayiinde %X, ... sektöründe %Y olarak gerçekleşmiştir. ... itibarıyla
... sektöründe kapasite kullanım oranı %Z düzeyinde gerçekleşmiştir.")

[SEKIL5]
ÜFE paragrafı. Sektör ÜFE'sini imalat ÜFE'siyle karşılaştır; varsa reel ciro
etkisine değin. (Örnek üslup: "... ÜFE'si imalat ÜFE'si ile paralel seyretmektedir.
... ayında söz konusu alt endeks değişimi %X, imalat ÜFE değişimi ise %Y düzeyindedir.")
{f'''
[SEKIL6]
İSO 500/İkinci 500 kurumsal görünüm paragrafı. Sektörden listeye giren firma sayısı,
toplam üretimden satışlar ve ihracat büyüklüğü, kârlılık ve yoğunlaşmayı kurumsal
yatırımcı bakışıyla değerlendir.
''' if iso_agg else ''}
Sadece istenen etiketleri döndür. Madde işareti ve markdown yıldızı (**) yasak."""

    # Yanit GERCEK etiket satirlari + dolu Turkce icerik icermeli.
    # (reasoning modellerin Ingilizce dusunce dokumu / etiketsiz cikti elenir)
    def _valid(t):
        if len(t) < 400:
            return False
        # kendi satirinda [GIRIS]/[SEKILn] etiketleri
        tags = re.findall(r'^\s*\[(GIRIS|SEKIL\d)\]\s*$', t, re.MULTILINE)
        if len(set(tags)) < 4:
            return False
        # Ingilizce reasoning sizintisi belirtileri
        low = t[:200].lower()
        if any(k in low for k in ("we need to", "let me", "okay,", "sure,", "here is", "i will")):
            return False
        # etiketlerden sonra dolu icerik say
        parts = re.split(r'^\s*\[[A-Z0-9]+\]\s*$', t, flags=re.MULTILINE)
        filled = sum(1 for p in parts if len(p.strip()) > 40)
        return filled >= 4
    try:
        # Tum key'ler denensin (retries=0 -> siniersiz), calisani bulana kadar hizlica
        return call_llm(prompt, system=SYSTEM_PROMPT,
                        max_tokens=(1200 if kisa else 2000), validate=_valid, retries=0)
    except Exception as e:
        # Hicbir key calismazsa: gercek verilerden deterministik rapor uret (bos kalmasin)
        print(f"   [LLM] tum key'ler basarisiz ({e}); deterministik rapora geciliyor.")
        return fallback_report(nace, fig1, fig2, fig3, fig4, fig5,
                               iso_agg=iso_agg, fig6=fig6, fig7=fig7)


# ─── EXCEL 'Analiz' sayfasi (eski uyumluluk icin tutuluyor) ──────────────────
def add_analysis_sheet(wb, nace, analysis_text):
    """
    Analiz metnini Excel workbook'a 'Analiz' sayfasi olarak ekler.
    (generate_report.py geriye donuk uyumluluk icin kullanir.)
    """
    from openpyxl.styles import Font, Alignment, PatternFill
    sector = SECTOR_NAMES.get(nace, nace)

    ws = wb.create_sheet('Analiz', 1)
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 100
    ws.column_dimensions['C'].width = 2

    H_FILL = PatternFill('solid', fgColor='44546A')
    S_FILL = PatternFill('solid', fgColor='00538F')

    def title_row(row, text, fill, size=13):
        ws.merge_cells(f'B{row}:B{row}')
        c = ws.cell(row=row, column=2, value=text)
        c.fill = fill
        c.font = Font(name='Arial', bold=True, size=size, color='FFFFFF')
        c.alignment = Alignment(horizontal='left', vertical='center',
                                wrap_text=True, indent=1)
        ws.row_dimensions[row].height = 28
        return row + 1

    def text_row(row, text, bold=False):
        ws.merge_cells(f'B{row}:B{row}')
        c = ws.cell(row=row, column=2, value=text)
        c.font = Font(name='Arial', bold=bold, size=9)
        c.alignment = Alignment(horizontal='left', vertical='top',
                                wrap_text=True, indent=1)
        return row + 1

    r = 2
    r = title_row(r, 'SEKTÖREL ANALİZ RAPORU', H_FILL, 14)
    r = title_row(r, f'{nace} — {sector}', S_FILL, 12)
    r += 1

    if analysis_text:
        for line in analysis_text.splitlines():
            line = line.strip()
            if re.match(r'^\[[A-Z0-9]+\]$', line):
                r += 1
                r = text_row(r, line.strip('[]'), bold=True)
            elif line:
                r = text_row(r, line)
            else:
                r += 1
    return wb
