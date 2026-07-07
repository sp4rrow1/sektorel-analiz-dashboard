# -*- coding: utf-8 -*-
"""
NACE Kodu -> Excel Raporu Uretici
Kullanim:
  python -X utf8 generate_report.py C13
  python -X utf8 generate_report.py C14
  python -X utf8 generate_report.py C29
  python -X utf8 generate_report.py C20

Onbellek yoksa otomatik olarak cache_all.py calistirilir.
"""
import pickle, os, sys, subprocess
from datetime import datetime
from nace_config import (SITC_MAP, SECTOR_NAMES, get_kko_code,
                         ALL_MANUFACTURING, SECTOR_SELECT_OPTIONS, TOTAL_MANUFACTURING)

CACHE_FILE = os.path.join(os.path.dirname(__file__), "data_cache.pkl")

MAX_ALT_C_SERIES  = 6   # Sekil 1'de gosterilecek max 3-hane grup
MAX_SINIF_SERIES  = 6   # Sekil 2'de gosterilecek max 4-hane sinif

# ─── CACHE ───────────────────────────────────────────────────────────────────
def load_cache():
    if not os.path.exists(CACHE_FILE):
        print("Onbellek bulunamadi, veriler cekiliyor (ilk seferlik ~10 dk)...")
        script = os.path.join(os.path.dirname(__file__), "cache_all.py")
        subprocess.run([sys.executable, "-X", "utf8", script], check=True)
    with open(CACHE_FILE, 'rb') as f:
        return pickle.load(f)

# ─── HESAPLAMALAR ────────────────────────────────────────────────────────────
def yoy_from_index(idx):
    result = {}
    for p, v in sorted(idx.items()):
        try:
            yr, mo = p.split('-')
            prev = f"{int(yr)-1}-{mo}"
        except ValueError:
            continue
        if prev in idx and idx[prev] != 0:
            result[p] = round((v / idx[prev] - 1) * 100, 4)
    return result

# ─── SEKIL 1: Uretim Endeksi (YoY %) ─────────────────────────────────────────
def build_sekil1(nace, alt_c_series, ana_c_series=None):
    """
    ALT_C'den verilen NACE'in dogrudan alt gruplarini (bir sonraki hane) bulur.
    C13 -> C131, C132, C133, C139
    C20 -> C201, C202, C203, C204, C205, C206, C207, C208

    nace == 'C' (ana imalat sanayii toplami) icin ozel yol: ANA_C
    dataflow'undan dogrudan Turkiye imalat sanayii genel endeksi (tek seri).
    """
    if nace == 'C':
        if not ana_c_series:
            return {}
        cands = [s for s in ana_c_series
                 if s['key'].get('ACTIVITY_NACE_REV2') == 'C'
                 and s['key'].get('DEGISIM_TUR') == '2'
                 and s['key'].get('SEASONAL_ADJUST') == 'Y']
        if cands:
            best = max(cands, key=lambda x: len(x['data']))
            return {'İmalat Sanayii Geneli': best['data']} if best['data'] else {}
        # Fallback: endeksten hesapla
        idx_cands = [s for s in ana_c_series
                     if s['key'].get('ACTIVITY_NACE_REV2') == 'C'
                     and s['key'].get('DEGISIM_TUR') == '1']
        w = [s for s in idx_cands if s['key'].get('SEASONAL_ADJUST') == 'W']
        idx_cands = w or idx_cands
        if not idx_cands: return {}
        best = max(idx_cands, key=lambda x: len(x['data']))
        y = yoy_from_index(best['data'])
        return {'İmalat Sanayii Geneli': y} if y else {}

    depth = len(nace) + 1  # C13 (3 hane) -> C13x (4 hane)

    target = {s['key'].get('ACTIVITY_NACE_REV2','') for s in alt_c_series
              if s['key'].get('ACTIVITY_NACE_REV2','').startswith(nace)
              and len(s['key'].get('ACTIVITY_NACE_REV2','')) == depth}

    if not target:
        return {}

    yayim = sorted({s['key'].get('YAYIM_DONEMI','') for s in alt_c_series
                    if s['key'].get('ACTIVITY_NACE_REV2') in target})
    latest = yayim[-1] if yayim else ''

    # DEGISIM_TUR=2 (YoY), SEASONAL_ADJUST=Y tercih et
    result = {}
    for nace_sub in sorted(target):
        # Oncelik: YoY direkt mevcut
        cands = [s for s in alt_c_series
                 if s['key'].get('ACTIVITY_NACE_REV2') == nace_sub
                 and s['key'].get('DEGISIM_TUR') == '2'
                 and s['key'].get('SEASONAL_ADJUST') == 'Y'
                 and s['key'].get('YAYIM_DONEMI') == latest]
        if not cands:
            # Fallback: endeksten hesapla
            idx_cands = [s for s in alt_c_series
                         if s['key'].get('ACTIVITY_NACE_REV2') == nace_sub
                         and s['key'].get('DEGISIM_TUR') == '1'
                         and s['key'].get('YAYIM_DONEMI') == latest]
            w = [s for s in idx_cands if s['key'].get('SEASONAL_ADJUST') == 'W']
            idx_cands = w or idx_cands
            if not idx_cands: continue
            best = max(idx_cands, key=lambda x: len(x['data']))
            y = yoy_from_index(best['data'])
            if y: result[nace_sub] = y
        else:
            best = max(cands, key=lambda x: len(x['data']))
            if best['data']: result[nace_sub] = best['data']

    # Cok fazla seri varsa en uzun olanlari sec
    if len(result) > MAX_ALT_C_SERIES:
        result = dict(sorted(result.items(), key=lambda x: len(x[1]), reverse=True)[:MAX_ALT_C_SERIES])

    return result

# ─── SEKIL 2: Alt Kirilimlar / Sinif Duzeyi (YoY %) ──────────────────────────
def build_sekil2(nace, sinif_series, ana_c_series=None):
    """
    SINIF_O'dan verilen NACE'in sinif duzeyindeki serilerini bulur.
    C13 -> C1311, C1312, C1391, C1392, ...

    nace == 'C' icin ozel yol: ANA_C'den 24 alt sektorun (C10..C33) kendi
    genel endeksleri dogrudan cekilir — imalat sanayiinin sektorel kirilimi.
    """
    if nace == 'C':
        if not ana_c_series:
            return {}
        result = {}
        for code in ALL_MANUFACTURING:
            cands = [s for s in ana_c_series
                     if s['key'].get('ACTIVITY_NACE_REV2') == code
                     and s['key'].get('DEGISIM_TUR') == '2'
                     and s['key'].get('SEASONAL_ADJUST') == 'Y']
            if cands:
                best = max(cands, key=lambda x: len(x['data']))
                if best['data']: result[code] = best['data']
        return result

    depth = len(nace) + 2  # C13 (3 hane) -> C13xx (5 hane)

    target = {s['key'].get('ACTIVITY_NACE_REV2','') for s in sinif_series
              if s['key'].get('ACTIVITY_NACE_REV2','').startswith(nace)
              and len(s['key'].get('ACTIVITY_NACE_REV2','')) == depth}

    if not target:
        return {}

    yayim = sorted({s['key'].get('YAYIM_DONEMI','') for s in sinif_series
                    if s['key'].get('ACTIVITY_NACE_REV2') in target})
    latest = yayim[-1] if yayim else ''

    result = {}
    for nace_sub in sorted(target):
        cands = [s for s in sinif_series
                 if s['key'].get('ACTIVITY_NACE_REV2') == nace_sub
                 and s['key'].get('DEGISIM_TUR') == '1'
                 and s['key'].get('YAYIM_DONEMI') == latest]
        w = [s for s in cands if s['key'].get('SEASONAL_ADJUST') == 'W']
        chosen = w or cands
        if not chosen: continue
        best = max(chosen, key=lambda x: len(x['data']))
        y = yoy_from_index(best['data'])
        if y: result[nace_sub] = y

    if len(result) > MAX_SINIF_SERIES:
        result = dict(sorted(result.items(), key=lambda x: len(x[1]), reverse=True)[:MAX_SINIF_SERIES])

    return result

# ─── SEKIL 3: Dis Ticaret (YoY %) ────────────────────────────────────────────
def build_sekil3(nace, dis_ticaret):
    sitc = SITC_MAP.get(nace, 'T')
    result = {}
    for label, key in [("Ihracat YoY %", "ihracat"), ("Ithalat YoY %", "ithalat")]:
        series = dis_ticaret.get(key, [])
        cands = [s for s in series
                 if s['key'].get('DTE_SEKTOR_KODLARI') == sitc
                 and s['key'].get('ENDEKS_DEGISIM') == '1']
        if not cands:
            cands = [s for s in series
                     if s['key'].get('DTE_SEKTOR_KODLARI') == 'T'
                     and s['key'].get('ENDEKS_DEGISIM') == '1']
        if not cands: continue
        best = max(cands, key=lambda x: len(x['data']))
        y = yoy_from_index(best['data'])
        if y: result[label] = y
    return result

# ─── SEKIL 4: KKO (%) ─────────────────────────────────────────────────────────
def build_sekil4(nace, kko):
    result = {}
    if nace in kko and kko[nace]:
        lbl = f"{nace} - {SECTOR_NAMES.get(nace, nace)} KKO %"
        result[lbl] = kko[nace]
    if 'TOPLAM' in kko:
        result['Imalat Sanayii KKO %'] = kko['TOPLAM']
    return result

# ─── SEKIL 5: UFE (Yillik %) ──────────────────────────────────────────────────
def build_sekil5(nace, ufe_series):
    yoy_s = [s for s in ufe_series if s['key'].get('DEGISIM') == '4']

    result = {}

    if nace == 'C':
        # Ana imalat sanayii: sadece toplam UFE (C) tek seri olarak.
        priority = ['C']
    else:
        # nace ile baslayan tum UFE kodlari
        ufe_codes = sorted({s['key'].get('URUN_UFE_NACE_CPA','') for s in yoy_s
                            if s['key'].get('URUN_UFE_NACE_CPA','').startswith(nace)})
        # En kisa kod = genel (C13), sonra alt gruplar
        priority = sorted(ufe_codes, key=len)[:4]
        # Imalat toplam da ekle
        if 'C' not in priority:
            priority.append('C')

    for code in priority:
        m = [s for s in yoy_s if s['key'].get('URUN_UFE_NACE_CPA') == code]
        if m:
            best = max(m, key=lambda x: len(x['data']))
            lbl = f"UFE {code} Yillik %"
            result[lbl] = best['data']

    return result

# ─── SEKIL 6: Ciro Endeksi (YoY %) ──────────────────────────────────────────
def build_sekil6(nace, ciro_series):
    """
    DF_CIRO_ENDEKS_DEGISIM_C: ACTIVITY=nace, DEGISIM_TUR=2 (YoY), SEASONAL_ADJUST=Y, CIRO_GOSTERGE_TUR=_T
    Also include C (imalat toplam) for comparison.
    """
    result = {}
    yayim = sorted({s['key'].get('YAYIM_DONEMI', '') for s in ciro_series
                    if s['key'].get('ACTIVITY', '').startswith('C')})
    latest = yayim[-1] if yayim else ''

    for code in [nace, 'C']:
        cands = [s for s in ciro_series
                 if s['key'].get('ACTIVITY') == code
                 and s['key'].get('DEGISIM_TUR') == '2'
                 and s['key'].get('SEASONAL_ADJUST') == 'Y'
                 and s['key'].get('CIRO_GOSTERGE_TUR') == '_T'
                 and s['key'].get('YAYIM_DONEMI') == latest]
        if not cands:
            continue
        best = max(cands, key=lambda x: len(x['data']))
        lbl = f"Ciro YoY % — {SECTOR_NAMES.get(code, code)}" if code != 'C' else "Ciro YoY % — İmalat Toplam"
        if best['data']:
            result[lbl] = best['data']
    return result


# ─── SEKIL 7: Ücretli Çalışan (YoY %) ────────────────────────────────────────
def build_sekil7(nace, ucretli_series):
    """
    DF_UCRETLI_CALISAN_ISTATISTIKLERI_C: EKONOMIK_FAALIYET=nace, ENDEKS_DEGISIM=4 (YoY), SEASONAL_ADJUST=W
    Also include C (imalat toplam).
    """
    result = {}
    yayim = sorted({s['key'].get('YAYIM_DONEMI', '') for s in ucretli_series
                    if s['key'].get('EKONOMIK_FAALIYET', '').startswith('C')})
    latest = yayim[-1] if yayim else ''

    for code in [nace, 'C']:
        cands = [s for s in ucretli_series
                 if s['key'].get('EKONOMIK_FAALIYET') == code
                 and s['key'].get('ENDEKS_DEGISIM') == '4'
                 and s['key'].get('YAYIM_DONEMI') == latest]
        if cands:
            best = max(cands, key=lambda x: len(x['data']))
            lbl = f"Çalışan YoY % — {SECTOR_NAMES.get(code, code)}" if code != 'C' else "Çalışan YoY % — İmalat Toplam"
            if best['data']:
                result[lbl] = best['data']
            continue
        # fallback: ENDEKS_DEGISIM=2 (aylık %)
        cands2 = [s for s in ucretli_series
                  if s['key'].get('EKONOMIK_FAALIYET') == code
                  and s['key'].get('ENDEKS_DEGISIM') == '2'
                  and s['key'].get('SEASONAL_ADJUST') == 'Y'
                  and s['key'].get('YAYIM_DONEMI') == latest]
        if cands2:
            best2 = max(cands2, key=lambda x: len(x['data']))
            lbl = f"Çalışan Aylık % — {SECTOR_NAMES.get(code, code)}" if code != 'C' else "Çalışan Aylık % — İmalat"
            if best2['data']:
                result[lbl] = best2['data']
    return result
    return result

# ─── YENI SERILER (Fiyat, Saat, YD-UFE, TUFE) ──────────────────────────────
def build_dis_ticaret_fiyat(nace, dis_ticaret_fiyat):
    if not dis_ticaret_fiyat: return {}
    sitc = SITC_MAP.get(nace, 'T')
    result = {}
    for label, key in [("İhracat Fiyat YoY %", "ihracat_fiyat"), ("İthalat Fiyat YoY %", "ithalat_fiyat")]:
        series = dis_ticaret_fiyat.get(key, [])
        cands = [s for s in series if s['key'].get('DTE_SEKTOR_KODLARI') == sitc and s['key'].get('ENDEKS_DEGISIM') == '1']
        if not cands: cands = [s for s in series if s['key'].get('DTE_SEKTOR_KODLARI') == 'T' and s['key'].get('ENDEKS_DEGISIM') == '1']
        if cands: result[label] = max(cands, key=lambda x: len(x['data']))['data']
    return result

def build_saat_ucret(nace, cache_su):
    if not cache_su: return {}
    result = {}
    for label, key in [("Saat YoY %", "saat"), ("Brüt Maaş YoY %", "maas")]:
        series = cache_su.get(key, [])
        cands = [s for s in series if s['key'].get('EKONOMIK_FAALIYET') == nace and s['key'].get('ENDEKS_DEGISIM') == '4']
        if not cands: cands = [s for s in series if s['key'].get('EKONOMIK_FAALIYET') == 'C' and s['key'].get('ENDEKS_DEGISIM') == '4']
        if cands: result[label] = max(cands, key=lambda x: len(x['data']))['data']
    return result

def build_ydufe(nace, ydufe_series):
    if not ydufe_series: return {}
    cands = [s for s in ydufe_series if s['key'].get('URUN_UFE_NACE_CPA') == nace and s['key'].get('YDUFE_ENDK_DGS') == '3']
    if not cands: cands = [s for s in ydufe_series if s['key'].get('URUN_UFE_NACE_CPA') == 'C' and s['key'].get('YDUFE_ENDK_DGS') == '3']
    if cands: return {f"YD-ÜFE YoY % — {SECTOR_NAMES.get(nace, nace)}": max(cands, key=lambda x: len(x['data']))['data']}
    return {}

def build_tufe(tufe_series):
    if not tufe_series: return {}
    cands = [s for s in tufe_series if s['key'].get('HRCG') == '00' and s['key'].get('TUFE_ENDK_DGS') == '3']
    if cands: return {"TÜFE YoY % — Türkiye Geneli": max(cands, key=lambda x: len(x['data']))['data']}
    return {}
# ─── EXCEL ───────────────────────────────────────────────────────────────────
def build_excel(nace, fig1, fig2, fig3, fig4, fig5, out_dir=None, save=True):
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.chart import LineChart, BarChart, Reference
    from openpyxl.utils import get_column_letter

    sector_name = SECTOR_NAMES.get(nace, nace)
    if out_dir is None:
        out_dir = os.path.expanduser("~\\Desktop")
    out_file = os.path.join(out_dir, f"{nace}_{sector_name[:30].replace(' ','_')}.xlsx")

    def fill(c): return PatternFill("solid", fgColor=c)
    def fnt(bold=False, color="000000", size=10):
        return Font(name="Calibri", bold=bold, color=color, size=size)
    thin = Side(style="thin", color="B0C4DE")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    C    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    L    = Alignment(horizontal="left",   vertical="center")
    R    = Alignment(horizontal="right",  vertical="center")

    H_FILL = fill("1F4E79"); H_FNT = fnt(True, "FFFFFF", 11)
    S_FILL = fill("2E75B6"); S_FNT = fnt(True, "FFFFFF", 10)
    A_FILL = fill("D6E4F0")
    W_FILL = fill("FFFFFF")
    B_FNT  = fnt(True)
    N_FNT  = fnt()

    def hdr(ws, r, c, txt, f=None, fn=None):
        cl = ws.cell(row=r, column=c, value=txt)
        cl.fill = f or H_FILL; cl.font = fn or H_FNT
        cl.alignment = C; cl.border = bdr

    def cel(ws, r, c, val, f=None, fn=None, al=None, fmt=None):
        cl = ws.cell(row=r, column=c, value=val)
        cl.fill = f or W_FILL; cl.font = fn or N_FNT
        cl.alignment = al or R; cl.border = bdr
        if fmt: cl.number_format = fmt
        return cl

    def make_sheet(wb, tab, data_dict, title, source, ctype="line", color_cells=True):
        if not data_dict:
            return
        ws = wb.create_sheet(title=tab[:31])
        ws.freeze_panes = "B3"
        nc = len(data_dict)

        ws.merge_cells(f"A1:{get_column_letter(nc+1)}1")
        c = ws["A1"]
        c.value = title; c.fill = H_FILL
        c.font  = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
        c.alignment = C
        ws.row_dimensions[1].height = 30

        hdr(ws, 2, 1, "Donem", S_FILL, S_FNT)
        labels = list(data_dict.keys())
        for i, lbl in enumerate(labels, 2):
            hdr(ws, 2, i, lbl, S_FILL, S_FNT)
        ws.row_dimensions[2].height = 40
        ws.column_dimensions["A"].width = 12
        for i in range(2, nc+2):
            ws.column_dimensions[get_column_letter(i)].width = 22

        periods = sorted({p for d in data_dict.values() for p in d})
        for ri, p in enumerate(periods, 3):
            rf = A_FILL if ri % 2 == 0 else W_FILL
            cel(ws, ri, 1, p, rf, B_FNT, C)
            for ci, lbl in enumerate(labels, 2):
                v  = data_dict[lbl].get(p)
                cl = cel(ws, ri, ci, v, rf, fmt="0.00")
                if v is not None and color_cells:
                    cl.font = Font(name="Calibri", size=10,
                                   color="C00000" if v < 0 else "375623")

        sr = len(periods) + 3
        ws.merge_cells(f"A{sr}:{get_column_letter(nc+1)}{sr}")
        c = ws[f"A{sr}"]
        c.value = source
        c.font  = Font(name="Calibri", italic=True, size=9, color="595959")
        c.alignment = L

        nrows = len(periods)
        chart = LineChart() if ctype == "line" else BarChart()
        if ctype == "bar": chart.barDir = "col"; chart.grouping = "clustered"
        chart.style = 10; chart.y_axis.title = "(%)"
        chart.height = 14; chart.width = 26
        chart.set_categories(Reference(ws, min_col=1, min_row=3, max_row=2+nrows))
        for ci in range(2, nc+2):
            s = openpyxl.chart.Series(
                Reference(ws, min_col=ci, min_row=2, max_row=2+nrows),
                title_from_data=True)
            chart.series.append(s)
        ws.add_chart(chart, f"A{sr+2}")

    wb = Workbook()
    wb.remove(wb.active)

    # Bilgi sayfasi
    ws0 = wb.create_sheet("Bilgi", 0)
    ws0.column_dimensions["A"].width = 28
    ws0.column_dimensions["B"].width = 65
    ws0.merge_cells("A1:B1")
    c = ws0["A1"]
    c.value = f"TUIK + TCMB | {nace}: {sector_name}"
    c.fill = H_FILL
    c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=13)
    c.alignment = C
    ws0.row_dimensions[1].height = 30
    rows = [
        ("Olusturma Tarihi", datetime.now().strftime("%d.%m.%Y %H:%M")),
        ("NACE Kodu",        f"{nace} - {sector_name}"),
        ("Kaynak 1",         "TUIK SDMX REST API v1.5 | nsiws.tuik.gov.tr/rest"),
        ("Kaynak 2",         "TCMB EVDS | evds3.tcmb.gov.tr"),
        ("",""),
        ("Sekil 1", f"Uretim Endeksi Alt Gruplari (YoY %) | ALT_C"),
        ("Sekil 2", f"Sinif Duzeyi Kirilimlar (YoY %) | SINIF_O (Endeksten hesaplandi)"),
        ("Sekil 3", f"Dis Ticaret Endeksi (YoY %) | SITC {SITC_MAP.get(nace,'T')}"),
        ("Sekil 4", f"Kapasite Kullanim Orani (%) | TCMB EVDS "
                   f"{'TOPLAM' if nace == TOTAL_MANUFACTURING else get_kko_code(nace)}"),
        ("Sekil 5", f"UFE Yillik % | NACE CPA {nace} ve Imalat Toplam"),
    ]
    for i, (k, v) in enumerate(rows, 2):
        ws0.cell(row=i, column=1, value=k).font = B_FNT
        ws0.cell(row=i, column=2, value=v).font = N_FNT

    make_sheet(wb, "Sekil1-UretimEndeksi", fig1,
               f"Sekil 1: {sector_name} Uretim Endeksi (Onceki Yila Gore Degisim, %)",
               "Kaynak: TUIK - Sanayi Uretim Endeksi (Mevsim+Takvim Etk. Arindirilmis)",
               "line")

    make_sheet(wb, "Sekil2-AltKirilimlar", fig2,
               f"Sekil 2: {sector_name} Alt Kirilimlar (Onceki Yila Gore Degisim, %)",
               "Kaynak: TUIK - Sanayi Uretim Endeksi, Sinif Duzeyi",
               "bar")

    make_sheet(wb, "Sekil3-DisTicaret", fig3,
               f"Sekil 3: {sector_name} Dis Ticaret (Onceki Yila Gore Degisim, %)",
               f"Kaynak: TUIK - Dis Ticaret (Takvim Etk. Arindirilmis) | SITC {SITC_MAP.get(nace,'T')}",
               "bar")

    make_sheet(wb, "Sekil4-KKO", fig4,
               f"Sekil 4: Kapasite Kullanim Orani - Yilliklandirilmis (%)",
               "Kaynak: TCMB EVDS - Imalat Sanayi Kapasite Kullanim Orani (NACE Rev.2)",
               "line", color_cells=False)

    make_sheet(wb, "Sekil5-UFE", fig5,
               f"Sekil 5: UFE Degisimi - {sector_name} Yillik %",
               "Kaynak: TUIK - Uretici Fiyat Endeksi (UFE)",
               "line")

    if save:
        wb.save(out_file)
    return wb, out_file

# ─── ANA ──────────────────────────────────────────────────────────────────────
def generate(nace_code, out_dir=None, with_analysis=True):
    nace = nace_code.upper().strip()
    if nace not in SECTOR_SELECT_OPTIONS:
        print(f"HATA: '{nace}' tanimsiz NACE kodu.")
        print(f"Gecerli kodlar: {', '.join(SECTOR_SELECT_OPTIONS)}")
        return None

    print(f"\n{'='*60}")
    print(f"Rapor uretiliyor: {nace} - {SECTOR_NAMES.get(nace, nace)}")
    print('='*60)

    cache = load_cache()
    created = cache.get('meta', {}).get('created_at', 'bilinmiyor')
    print(f"Onbellek: {created[:10]}\n")

    print("-> Sekil 1: Uretim endeksi alt gruplari...")
    f1 = build_sekil1(nace, cache['alt_c'], ana_c_series=cache.get('ana_c'))
    print(f"   {len(f1)} seri")

    print("-> Sekil 2: Sinif duzeyinde alt kirilimlar...")
    f2 = build_sekil2(nace, cache['sinif_o'], ana_c_series=cache.get('ana_c'))
    print(f"   {len(f2)} seri")

    print("-> Sekil 3: Dis ticaret...")
    f3 = build_sekil3(nace, cache['dis_ticaret'])
    print(f"   {len(f3)} seri")

    print("-> Sekil 4: Kapasite kullanim orani...")
    f4 = build_sekil4(nace, cache['kko'])
    print(f"   {len(f4)} seri")

    print("-> Sekil 5: UFE...")
    f5 = build_sekil5(nace, cache['ufe'])
    print(f"   {len(f5)} seri")

    # Turev metrikler icin (reel ciro, verimlilik) — LLM raporunda kullanilir
    f6 = build_sekil6(nace, cache['ciro'])    if 'ciro'    in cache else {}
    f7 = build_sekil7(nace, cache['ucretli']) if 'ucretli' in cache else {}

    # ─── ISO 500 verisi ──────────────────────────────────────────────────────
    iso_agg = None
    try:
        from iso_data import sector_iso
        iso_agg = sector_iso(nace)
        if iso_agg:
            print(f"-> ISO 500: {iso_agg['firma_sayisi']} firma ({iso_agg['yil']})")
    except Exception as e:
        print(f"-> ISO verisi atlandi: {e}")

    # ─── PRODTR ürün verisi ──────────────────────────────────────────────────
    prodtr_agg = None
    try:
        from prodtr_data import sector_products
        prodtr_agg = sector_products(nace)
        if prodtr_agg:
            print(f"-> PRODTR: {prodtr_agg['urun_sayisi']} ürün ({prodtr_agg['yil']})")
    except Exception as e:
        print(f"-> PRODTR verisi atlandi: {e}")

    print("\n-> Excel olusturuluyor...")
    wb, out = build_excel(nace, f1, f2, f3, f4, f5, out_dir, save=False)

    # ─── LLM analizi ─────────────────────────────────────────────────────────
    analysis_text = None
    if with_analysis:
        print("\n-> LLM analizi uretiliyor...")
        try:
            from sector_analysis import generate_analysis, add_analysis_sheet
            analysis_text = generate_analysis(nace, f1, f2, f3, f4, f5, iso_agg=iso_agg,
                                               fig6=f6, fig7=f7)
            add_analysis_sheet(wb, nace, analysis_text)
            print("   Excel analiz sayfasi eklendi.")
        except Exception as e:
            print(f"   LLM analizi basarisiz (devam ediliyor): {e}")

    import openpyxl
    wb.save(out)
    print(f"\nExcel kaydedildi: {out}")

    # ─── Haber & risk analizi ────────────────────────────────────────────────
    news_text = None
    if with_analysis:
        try:
            from news_analysis import fetch_sector_news, analyze_news
            items = fetch_sector_news(nace)
            if items:
                print(f"\n-> Haber analizi: {len(items)} baslik sentezleniyor...")
                news_text = analyze_news(nace, items)
        except Exception as e:
            print(f"   Haber analizi atlandi: {e}")

    # ─── Word raporu (Kimteks formati) ───────────────────────────────────────
    print("\n-> Word raporu olusturuluyor (Kimteks formati)...")
    try:
        import tempfile
        from report_charts import generate_all_charts
        from report_word   import build_word_report

        tmpdir = tempfile.mkdtemp(prefix='tuik_charts_')
        chart_paths = generate_all_charts(nace, f1, f2, f3, f4, f5, tmpdir,
                                          iso_agg=iso_agg)
        print(f"   {len(chart_paths)} grafik olusturuldu.")

        word_file = build_word_report(
            nace, f1, f2, f3, f4, f5,
            analysis_text=analysis_text,
            chart_paths=chart_paths,
            out_dir=out_dir,
            iso_agg=iso_agg,
            news_analysis=news_text,
            prodtr_agg=prodtr_agg,
        )
        print(f"Word kaydedildi: {word_file}")

        # Gecici grafik dosyalarini sil
        import shutil
        shutil.rmtree(tmpdir, ignore_errors=True)
        return word_file

    except Exception as e:
        print(f"   Word raporu basarisiz: {e}")
        import traceback; traceback.print_exc()
        return out


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Kullanim: python -X utf8 generate_report.py <NACE_KODU> [--no-llm] [cikti_klasoru]")
        print(f"Ornek:    python -X utf8 generate_report.py C13")
        print(f"          python -X utf8 generate_report.py C29 --no-llm")
        print(f"Gecerli kodlar: {', '.join(ALL_MANUFACTURING)}")
        sys.exit(1)

    args = sys.argv[1:]
    no_llm   = "--no-llm" in args
    args     = [a for a in args if a != "--no-llm"]
    nace_arg = args[0]
    out_arg  = args[1] if len(args) > 1 else None
    generate(nace_arg, out_arg, with_analysis=not no_llm)
